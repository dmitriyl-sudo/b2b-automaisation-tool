# --- ОБЯЗАТЕЛЬНЫЕ ИМПОРТЫ ДЛЯ ВСЕГО ФАЙЛА ---
# Убедитесь, что все эти библиотеки установлены (pip install ...)

# Стандартные и сторонние библиотеки
import logging
import os
import pickle
import time
import json
import sys
import re
from datetime import timedelta, datetime
from typing import Literal, List, Dict, Optional, Any, Tuple

import pandas as pd
import openpyxl
from openpyxl import Workbook
from pydantic import BaseModel
from fastapi import FastAPI, HTTPException, BackgroundTasks, Body, Depends
from fastapi.responses import JSONResponse
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import jwt, JWTError

# Импорты для работы с Google API
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

# --- ВАЖНО: Импорты из вашего проекта ---
from main import geo_groups, password_data, site_list, GLITCHSPIN_EXTRA_GEOS
from utils.excel_utils import save_payment_data_to_excel, merge_payment_data
from utils.google_drive import create_google_file, upload_table_to_sheets, get_credentials
from utils.google_drive import finalize_google_sheet_formatting

from extractors.ritzo_extractor import RitzoExtractor
from extractors.rolling_extractor import RollingExtractor
from extractors.nfs_extractor import NeedForSpinExtractor
from extractors.wld_extractor import WildTokyoExtractor
from extractors.godofwins_extractor import GodofwinsExtractor
from extractors.hugo_extractor import HugoExtractor
from extractors.winshark_extractor import WinsharkExtractor
from extractors.spinlander_extractor import SpinlanderExtractor
from extractors.slota_extractor import SlotaExtractor
from extractors.spinline_extractor import SpinlineExtractor
from extractors.glitchspin_extractor import GlitchSpinExtractor

from test_runner import run_payment_method_tests


# --- КОНФИГУРАЦИЯ ЛОГГИРОВАНИЯ ---
LOG_FILE = os.path.join(os.path.dirname(__file__), "export_debug.log")
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, mode='w', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)

# --- БЕЗОПАСНОСТЬ И АУТЕНТИФИКАЦИЯ ---
SECRET_KEY = "supersecretkey123"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60

users_db = {
    "admin": {"password": "123", "role": "admin"},
    "qa": {"password": "qa123", "role": "qa"},
    "viewer": {"password": "viewer123", "role": "viewer"},
}

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="login")

def create_access_token(data: dict, expires_delta: timedelta = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=15))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def get_current_user(token: str = Depends(oauth2_scheme)):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None or username not in users_db:
            raise HTTPException(status_code=401, detail="Invalid token")
        return {"username": username, "role": users_db[username]["role"]}
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

def extract_tags(name):
    tags = []
    dep_match = None
    for i in range(10):
        if f"{i}dep" in name.lower():
            dep_match = f"{i}DEP"
            break
    if dep_match:
        tags.append(dep_match)
    if "aff" in name.lower():
        tags.append("AFF")
    if "mob" in name.lower():
        tags.append("MOB")
    return tags

def extract_conditions_from_name(name: str) -> str:
    if not isinstance(name, str):
        return "ALL"
    tags = set()
    name_lower = name.lower()
    dep_match = re.search(r'(\d+)dep', name_lower)
    if dep_match:
        tags.add(f"{dep_match.group(1)}DEP")
    if "aff" in name_lower:
        tags.add("AFF")
    if "mob" in name_lower:
        tags.add("MOB")
    if not tags:
        return "ALL"
    return "\n".join(sorted(list(tags)))


# --- ИНИЦИАЛИЗАЦИЯ FastAPI ---
app = FastAPI()


# --- PYDANTIC МОДЕЛИ ДЛЯ ЗАПРОСОВ ---
class LoginTestRequest(BaseModel):
    project: str
    login: Optional[str] = None
    geo: str
    env: Literal["stage", "prod"]
    mode: Optional[str] = None

class MethodTestRequest(BaseModel):
    project: str
    geo: str
    login: str
    mode: str
    env: Literal["stage", "prod"]

class FullProjectExportRequest(BaseModel):
    project: str
    env: Literal["stage", "prod"]


# --- ГЛОБАЛЬНАЯ КОНФИГУРАЦИЯ ---
EXTRACTORS = {
    site["name"]: (site["extractor_class"], site["stage_url"], site["prod_url"])
    for site in site_list
}

# 1) ПОМОЩНИКИ
def _key_join(title: str, name: str) -> str:
    return f"{(title or '').strip()}|||{(name or '').strip()}"

def _extract_pairs_and_minlist(methods_or_titles, names_opt=None):
    """
    Унифицирует оба формата:
      - старый: methods_or_titles = List[str], names_opt = List[str]
      - новый:  methods_or_titles = List[dict] с полями title/name/min_deposit
    Возвращает:
      pairs:    List[Tuple[str, str]]
      min_list: List[Dict[str, Any]]  # [{"title","name","min_deposit","currency"?, "min_source"?, ...}]
    """
    pairs, min_list = [], []

    if isinstance(methods_or_titles, list) and methods_or_titles and isinstance(methods_or_titles[0], dict):
        for it in methods_or_titles:
            t = (it.get("title") or it.get("alias") or it.get("name") or "").strip()
            n = (it.get("name") or it.get("alias") or it.get("doc_id") or "").strip()
            if not t or not n:
                continue
            pairs.append((t, n))
            if it.get("min_deposit") is not None:
                try:
                    row = {
                        "title": t,
                        "name": n,
                        "min_deposit": float(it["min_deposit"]),
                    }
                    if it.get("currency"):    row["currency"] = it["currency"]
                    if it.get("min_source"):  row["min_source"] = it["min_source"]
                    min_list.append(row)
                except (ValueError, TypeError):
                    continue
    else:
        if not isinstance(names_opt, list):
            names_opt = []
        pairs = list(zip(methods_or_titles or [], names_opt))
    return pairs, min_list
    
def _is_geo_forbidden_for_project(project: str, geo: str) -> bool:
    """
    Если у класса экстрактора есть статический метод is_geo_forbidden_static(geo),
    используем его, чтобы решить — пропускать GEO или нет.
    """
    try:
        extractor_class, _, _ = EXTRACTORS[project]
    except Exception:
        return False

    # Статический метод на классе
    if hasattr(extractor_class, "is_geo_forbidden_static"):
        try:
            return bool(extractor_class.is_geo_forbidden_static(geo))
        except Exception:
            return False

    # На всякий — инстанс и инстансовый метод
    try:
        tmp = extractor_class(login="", password="", base_url="")
        if hasattr(tmp, "is_geo_forbidden"):
            return bool(tmp.is_geo_forbidden(geo))
    except Exception:
        pass
    return False

# 2) GET_METHODS_ONLY
def get_methods_only(project: str, geo: str, env: str, login: str):
    if project not in EXTRACTORS:
        raise HTTPException(status_code=400, detail="Unknown project")
    extractor_class, stage_url, prod_url = EXTRACTORS[project]
    url = stage_url if env == "stage" else prod_url
    user_agent = (
        "Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1"
        if "mobi" in login
        else "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0.3 Safari/605.1.15"
    )
    extractor = extractor_class(login, password_data, user_agent=user_agent, base_url=url)
    if not extractor.authenticate():
        return {"success": False, "error": "Authentication failed"}
    try:
        deposit, withdraw, dep_names, wd_names, currency, dep_count, recommended = extractor.get_payment_and_withdraw_systems(geo)

        # Нормализуем оба формата
        deposit_pairs, dep_min_list = _extract_pairs_and_minlist(deposit, dep_names)
        withdraw_pairs, wd_min_list = _extract_pairs_and_minlist(withdraw, wd_names)
        min_deposit_list = dep_min_list + wd_min_list

        # Для фронта: мапа ключ->значение и легаси-массив
        min_deposit_by_key = {
            _key_join(x["title"], x["name"]): x["min_deposit"] for x in min_deposit_list
        }
        min_deposits_legacy = [
            {
                "Title": x.get("title", ""),
                "Name": x.get("name", ""),
                "MinDeposit": x.get("min_deposit", None),
                "Currency": x.get("currency", "") or currency or extractor.currency or ""
            }
            for x in min_deposit_list
        ]

        # Нормализация recommended + дедупликация по порядку
        # recommended приходит set[(title,name)], нормализуем пробелы
        recommended_norm = {(t.strip(), n.strip()) for (t, n) in (recommended or set())}
        recommended_pairs: List[Tuple[str, str]] = []
        seen_rec: set = set()
        for pair in (deposit_pairs + withdraw_pairs):
            p = (pair[0].strip(), pair[1].strip())
            if p in recommended_norm and p not in seen_rec:
                recommended_pairs.append(pair)
                seen_rec.add(p)

        return {
            "success": True,
            "currency": currency or extractor.currency,
            "deposit_methods": deposit_pairs,
            "withdraw_methods": withdraw_pairs,
            "recommended_methods": recommended_pairs,
            "min_deposit_map": min_deposit_list,       # новый формат (list of dicts)
            "min_deposit_by_key": min_deposit_by_key,  # быстрый доступ по ключу
            "min_deposits": min_deposits_legacy,       # ЛЕГАСИ для старого фронта
            "debug": {
                "recommended_pairs": list(recommended_norm),
                "deposit_raw_type": type(deposit).__name__,
                "withdraw_raw_type": type(withdraw).__name__,
                "deposit_len": len(deposit) if isinstance(deposit, list) else None,
                "withdraw_len": len(withdraw) if isinstance(withdraw, list) else None
            }
        }
    except Exception as e:
        logging.exception(f"Error in get_methods_only for {project}/{geo}/{login}: {e}")
        return {"success": False, "error": str(e)}

def run_login_check(project: str, geo: str, env: str, login: str):
    if project not in EXTRACTORS:
        raise HTTPException(status_code=400, detail="Unknown project")
    extractor_class, stage_url, prod_url = EXTRACTORS[project]
    url = stage_url if env == "stage" else prod_url
    user_agent = (
        "Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1"
        if "mobi" in login
        else "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0.3 Safari/605.1.15"
    )
    extractor = extractor_class(login, password_data, user_agent=user_agent, base_url=url)
    if not extractor.authenticate():
        raise HTTPException(status_code=401, detail="Authentication failed")
    return {
        "success": True,
        "currency": extractor.currency,
        "deposit_count": extractor.deposit_count
    }

# --- API ЭНДПОИНТЫ ---

@app.post("/login")
def login(form_data: OAuth2PasswordRequestForm = Depends()):
    user = users_db.get(form_data.username)
    if not user or user["password"] != form_data.password:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    access_token = create_access_token(data={"sub": form_data.username})
    return {"access_token": access_token, "token_type": "bearer"}

@app.get("/me")
def read_me(user: dict = Depends(get_current_user)):
    return user

@app.get("/list-projects")
def list_projects():
    return [
        {"name": site["name"], "stage_url": site["stage_url"], "prod_url": site["prod_url"]}
        for site in site_list
    ]

@app.get("/geo-groups")
def get_geo_groups():
    # Возвращаем базовые GEO + доп. GEO для Glitchspin, чтобы фронт их видел
    merged = {**geo_groups, **GLITCHSPIN_EXTRA_GEOS}
    return merged

@app.post("/get-methods-only")
def get_methods_only_endpoint(request: LoginTestRequest):
    # ⬇️ новое:
    if _is_geo_forbidden_for_project(request.project, request.geo):
        return {
            "success": True,
            "deposit_methods": [],
            "withdraw_methods": [],
            "recommended_methods": [],
            "skipped_geo": True,
            "reason": "forbidden_geo"
        }

    return get_methods_only(project=request.project, geo=request.geo, env=request.env, login=request.login)

# ЭНДПОИНТ ДЛЯ ПОЛУЧЕНИЯ МИНИМАЛЬНЫХ ДЕПОЗИТОВ
@app.post("/get-min-deposits")
def get_min_deposits_endpoint(request: LoginTestRequest):
    res = get_methods_only(project=request.project, geo=request.geo, env=request.env, login=request.login)
    if not res.get("success"):
        raise HTTPException(status_code=400, detail=res.get("error", "Unknown error"))
    # отдадим оба, чтобы фронт мог выбрать любой вариант
    return {
        "success": True,
        "min_deposit_map": res.get("min_deposit_map", []),
        "min_deposits": res.get("min_deposits", []),
        "min_deposit_by_key": res.get("min_deposit_by_key", {})
    }

@app.post("/run-login-check")
def run_login_check_endpoint(request: LoginTestRequest):
    # ⬇️ новое:
    if _is_geo_forbidden_for_project(request.project, request.geo):
        return {
            "success": True,
            "currency": None,
            "deposit_count": 0,
            "skipped_geo": True,
        }

    return run_login_check(project=request.project, geo=request.geo, env=request.env, login=request.login)

@app.post("/run-multi-auth-check")
def run_multi_auth_check_endpoint(request: LoginTestRequest):
    """
    Проверяет авторизацию для всех логинов в выбранном GEO
    """
    if request.mode != "geo":
        raise HTTPException(status_code=400, detail="Only 'geo' mode supported")
    
    if _is_geo_forbidden_for_project(request.project, request.geo):
        return {"results": []}
    
    if request.project not in EXTRACTORS:
        raise HTTPException(status_code=400, detail="Unknown project")
    
    # Получаем логины для GEO
    effective_geo_groups = geo_groups
    if request.project == "Glitchspin":
        effective_geo_groups = {**geo_groups, **GLITCHSPIN_EXTRA_GEOS}
    
    logins = effective_geo_groups.get(request.geo, [])
    if not logins:
        return {"results": []}
    
    results = []
    extractor_class, stage_url, prod_url = EXTRACTORS[request.project]
    url = stage_url if request.env == "stage" else prod_url
    
    for login in logins:
        try:
            user_agent = (
                "Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1"
                if "mobi" in login
                else "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0.3 Safari/605.1.15"
            )
            extractor = extractor_class(login, password_data, user_agent=user_agent, base_url=url)
            
            if extractor.authenticate():
                results.append({
                    "login": login,
                    "success": True,
                    "currency": extractor.currency,
                    "deposit_count": extractor.deposit_count
                })
            else:
                results.append({
                    "login": login,
                    "success": False,
                    "currency": None,
                    "deposit_count": 0
                })
        except Exception as e:
            results.append({
                "login": login,
                "success": False,
                "currency": None,
                "deposit_count": 0,
                "error": str(e)
            })
    
    return {"results": results}


@app.post("/test-methods")
async def test_methods(req: MethodTestRequest):
    results = run_payment_method_tests(
        project=req.project,
        geo=req.geo,
        login=req.login,
        mode=req.mode,
        env=req.env
    )
    return JSONResponse(content={"results": results})

@app.post("/export-table-to-sheets")
def export_table_to_sheets(payload: Dict = Body(...)):
    data: List[Dict] = payload.get("data", [])
    original_order: Optional[List[str]] = payload.get("originalOrder")
    try:
        file_id = upload_table_to_sheets(data, original_order=original_order)
        return {
            "success": True,
            "message": "Таблица экспортирована в Google Sheets",
            "sheet_url": f"https://docs.google.com/spreadsheets/d/{file_id}"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
        
# ЭКСПОРТ FULL-ПРОЕКТА В ЛОКАЛЬНЫЙ EXCEL
@app.post("/export-full-project")
def export_full_project(req: FullProjectExportRequest, background_tasks: BackgroundTasks):
    if req.project not in [site["name"] for site in site_list]:
        raise HTTPException(status_code=400, detail="Unknown project")

    extractor_class, stage_url, prod_url = EXTRACTORS[req.project]
    url = stage_url if req.env == "stage" else prod_url

    filename = f"merged_{req.project}_{req.env}_FULL.xlsx"
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    for geo, logins in geo_groups.items():
        # ⬇️ новое:
        if _is_geo_forbidden_for_project(req.project, geo):
            logging.info(f"Skipping export for project '{req.project}' and forbidden geo '{geo}'")
            continue
        
        if not logins:
            continue

        all_payment_data = {}
        original_order = []
        method_type_map = {}
        recommended_set = set()
        seen_titles = set()
        currency = None
        conditions_raw = {} 

        min_by_title: Dict[str, float] = {}
        had_any_min = False

        for login in logins:
            user_agent = (
                "Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X)"
                if "mobi" in login
                else "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"
            )
            extractor = extractor_class(login, password_data, user_agent=user_agent, base_url=url)
            if not extractor.authenticate():
                continue

            deposit, withdraw, dep_names, wd_names, curr, _, recommended = extractor.get_payment_and_withdraw_systems(geo)
            currency = currency or curr

            deposit_pairs, dep_min_list = _extract_pairs_and_minlist(deposit, dep_names)
            withdraw_pairs, wd_min_list = _extract_pairs_and_minlist(withdraw, wd_names)
            
            if dep_min_list:
                had_any_min = True
                for row in dep_min_list:
                    title = row.get("title")
                    val = row.get("min_deposit")
                    if title is not None and val is not None:
                        try:
                            val_float = float(val)
                            if title not in min_by_title:
                                min_by_title[title] = val_float
                            else:
                                min_by_title[title] = min(min_by_title[title], val_float)
                        except (ValueError, TypeError):
                            continue

            all_pairs = deposit_pairs + withdraw_pairs

            for title, name in all_pairs:
                key = f"{title}|||{name}"
                if title not in seen_titles:
                    original_order.append(title)
                    seen_titles.add(title)

                method_type_map[key] = method_type_map.get(key, {"Deposit": "NO", "Withdraw": "NO"})
                if (title, name) in deposit_pairs:
                    method_type_map[key]["Deposit"] = "YES"
                if (title, name) in withdraw_pairs:
                    method_type_map[key]["Withdraw"] = "YES"

                all_payment_data.setdefault(login, {})[key] = {
                    "Payment Name": name,
                    "Deposit": method_type_map[key]["Deposit"],
                    "Withdraw": method_type_map[key]["Withdraw"],
                    "Status": req.env.upper()
                }

                tags = extract_tags(name)
                if tags:
                    combined = "+".join(tags)
                    if title not in conditions_raw:
                        conditions_raw[title] = set()
                    conditions_raw[title].add(combined)
                else:
                    if title not in conditions_raw:
                        conditions_raw[title] = set()
                    conditions_raw[title].add("ALL")

            recommended_set.update(recommended)

        conditions_map = {k: "\n".join(sorted(v)) if v else "ALL" for k, v in conditions_raw.items()}
        merged = merge_payment_data(
            all_payment_data,
            list(all_payment_data.keys()),
            original_order,
            currency,
            list(recommended_set),
            excel_filename=filename,
            url=url,
            conditions_map=conditions_map
        )

        sheet = wb.create_sheet(title=f"{req.project}_{geo}_{req.env}")
        headers = ["Paymethod", "Payment Name", "Currency", "Deposit", "Withdraw", "Status", "Min Deposit", "Details"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        for i, method in enumerate(merged, start=2):
            row_min = min_by_title.get(method) if had_any_min else None
            row_data = [
                method, 
                merged[method].get("Payment Name", ""),
                merged[method].get("Currency", ""),
                merged[method].get("Deposit", ""),
                merged[method].get("Withdraw", ""),
                merged[method].get("Status", ""),
                row_min if row_min is not None else "",
                merged[method].get("Details", "")
            ]
            for col, val in enumerate(row_data, start=1):
                sheet.cell(row=i, column=col, value=val)

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(filename)
    
    try:
        creds = get_credentials()
        file_metadata = {
            'name': filename,
            'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        media = MediaFileUpload(filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        drive_service = build('drive', 'v3', credentials=creds)
        
        file = drive_service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        file_id = file.get('id')
        
        os.remove(filename)

        return {
            "success": True,
            "message": "Проект экспортирован с несколькими GEO-листами и загружен на Google Drive",
            "sheet_url": f"https://docs.google.com/spreadsheets/d/{file_id}"
        }
    except Exception as e:
        logging.error(f"Failed to upload merged Excel file to Google Drive: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to upload file: {str(e)}")


# ЭКСПОРТ FULL-ПРОЕКТА В GOOGLE SHEETS
@app.post("/export-full-project-to-google-sheet")
def export_full_project_to_google_sheet(data: FullProjectExportRequest):
    """
    Экспортирует все GEO для проекта в один Google Sheet,
    где каждый GEO находится на отдельном листе.
    """
    try:
        creds = get_credentials()
        sheets_service = build("sheets", "v4", credentials=creds)
    except Exception as e:
        logging.error(f"Google API credentials error: {e}")
        raise HTTPException(status_code=500, detail=f"Google API credentials error: {e}")

    project = data.project
    env = data.env
    all_geo = list(geo_groups.keys())

    spreadsheet_body = { "properties": {"title": f"📊 {project} ({env.upper()}) — Full Export"} }
    try:
        spreadsheet = sheets_service.spreadsheets().create(body=spreadsheet_body).execute()
        file_id = spreadsheet["spreadsheetId"]
        logging.info(f"Created new Google Sheet with ID: {file_id}")
    except Exception as e:
        logging.error(f"Failed to create Google Sheet: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to create Google Sheet: {e}")

    sheets_to_add = []
    all_sheets_data = {}

    for geo in all_geo:
        # ⬇️ новое:
        if _is_geo_forbidden_for_project(project, geo):
            logging.info(f"Skipping sheet creation for project '{project}' and forbidden geo '{geo}'")
            continue
        
        logins = geo_groups.get(geo, [])
        if not logins:
            continue

        seen_methods = {}
        recommended_set = set()

        # Мапа для сбора минимальных депозитов по всем логинам в рамках одного GEO
        min_deposits_by_key_geo: Dict[str, Dict] = {}

        for login in logins:
            try:
                methods = get_methods_only(project=project, geo=geo, env=env, login=login)
                if not methods or not methods.get("success"):
                    continue
            except Exception as e:
                logging.warning(f"Could not fetch methods for {login} in {geo}: {e}")
                continue

            for (title, name) in methods.get("recommended_methods", []):
                recommended_set.add((title.strip(), name.strip()))

            deposit_pairs = methods.get("deposit_methods", [])
            withdraw_pairs = methods.get("withdraw_methods", [])
            all_pairs = set(deposit_pairs + withdraw_pairs)

            # Собираем минимумы для этого логина и объединяем в общую мапу по GEO
            min_deposits_map = { _key_join(d['title'], d['name']): d for d in methods.get('min_deposit_map', []) }
            for key, min_info in min_deposits_map.items():
                if key not in min_deposits_by_key_geo:
                    min_deposits_by_key_geo[key] = min_info
                else:
                    # Если найдено меньшее значение, обновляем
                    current_min = min_deposits_by_key_geo[key].get('min_deposit')
                    new_min = min_info.get('min_deposit')
                    if new_min is not None and (current_min is None or new_min < current_min):
                        min_deposits_by_key_geo[key]['min_deposit'] = new_min

            for (title, name) in all_pairs:
                key = (title.strip(), name.strip())
                if key not in seen_methods:
                    seen_methods[key] = {
                        "Paymethod": title,
                        "Recommended": "",
                        "Deposit": "NO",
                        "Withdraw": "NO",
                        "Payment Name": name,
                        "Min Dep": "", 
                        "Conditions": extract_conditions_from_name(name),
                        "Env": env.upper()
                    }
                
                if (title, name) in deposit_pairs:
                    seen_methods[key]["Deposit"] = "YES"
                if (title, name) in withdraw_pairs:
                    seen_methods[key]["Withdraw"] = "YES"
                if (title, name) in recommended_set:
                    seen_methods[key]["Recommended"] = "⭐"

        # Заполняем поле "Min Dep" после объединения всех логинов
        for key, method_data in seen_methods.items():
            min_info = min_deposits_by_key_geo.get(_key_join(key[0], key[1]))
            if min_info and min_info.get("min_deposit") is not None:
                method_data["Min Dep"] = f"{min_info['min_deposit']} {min_info.get('currency', '')}"

        if seen_methods:
            rows = list(seen_methods.values())
            headers = list(rows[0].keys())
            values = [headers] + [[r.get(h, "") for h in headers] for r in rows]
            
            sheet_title = geo[:100]
            sheets_to_add.append({"addSheet": {"properties": {"title": sheet_title}}})
            all_sheets_data[sheet_title] = values

    if not sheets_to_add:
        try:
            drive_service = build('drive', 'v3', credentials=creds)
            drive_service.files().delete(fileId=file_id).execute()
        except Exception as e:
            logging.error(f"Could not delete empty spreadsheet {file_id}: {e}")
        return {"success": False, "message": "No data found for any GEO in this project."}
        
    try:
        sheets_service.spreadsheets().batchUpdate(
            spreadsheetId=file_id,
            body={"requests": sheets_to_add}
        ).execute()

        data_to_update = []
        for sheet_title, values in all_sheets_data.items():
            data_to_update.append({
                "range": f"'{sheet_title}'!A1",
                "values": values
            })
        
        sheets_service.spreadsheets().values().batchUpdate(
            spreadsheetId=file_id,
            body={
                "valueInputOption": "RAW",
                "data": data_to_update
            }
        ).execute()
        
    except Exception as e:
        logging.error(f"Failed to update sheets in batch: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to update sheets: {e}")

    try:
        sheet_metadata = sheets_service.spreadsheets().get(spreadsheetId=file_id).execute()
        sheets = sheet_metadata.get('sheets', '')
        default_sheet_id = next((s['properties']['sheetId'] for s in sheets if s['properties']['title'] == 'Sheet1'), None)

        if default_sheet_id is not None:
             sheets_service.spreadsheets().batchUpdate(
                spreadsheetId=file_id,
                body={"requests": [{"deleteSheet": {"sheetId": default_sheet_id}}]}
            ).execute()
    except Exception:
        logging.warning("Could not delete default 'Sheet1'. It may not have existed.")

    return {"success": True, "sheet_url": f"https://docs.google.com/spreadsheets/d/{file_id}"}


# ЭКСПОРТ НЕСКОЛЬКИХ ТАБЛИЦ В GOOGLE SHEETS (ПО ФРОНТУ)
@app.post("/export-table-to-sheets-multi")
def export_table_to_sheets_multi(payload: Dict = Body(...)):
    """
    Принимает несколько таблиц по GEO и выгружает их в один Google Sheet (по фронту).
    """
    try:
        sheets = payload.get("sheets", [])
        if not sheets or not isinstance(sheets, list):
            raise HTTPException(status_code=400, detail="Invalid or missing 'sheets'")

        creds = get_credentials()
        sheets_service = build("sheets", "v4", credentials=creds)

        spreadsheet_body = {
            "properties": {
                "title": f"📊 Multi-GEO Export {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            }
        }
        spreadsheet = sheets_service.spreadsheets().create(body=spreadsheet_body).execute()
        file_id = spreadsheet["spreadsheetId"]
        logging.info(f"Created new spreadsheet: {file_id}")

        requests = []
        data_to_update = [] 

        for sheet_data_item in sheets:
            title = sheet_data_item.get("geo", "Sheet")[:100]
            rows = sheet_data_item.get("rows", [])
            if not rows:
                continue

            requests.append({"addSheet": {"properties": {"title": title}}})
            headers = list(rows[0].keys())
            values = [headers] + [[r.get(h, "") for h in headers] for r in rows]
            
            data_to_update.append({
                "range": f"'{title}'!A1",
                "values": values
            })

        if not requests:
            return {"success": False, "message": "No valid data to export"}

        sheets_service.spreadsheets().batchUpdate(
            spreadsheetId=file_id,
            body={"requests": requests}
        ).execute()

        sheets_service.spreadsheets().values().batchUpdate(
            spreadsheetId=file_id,
            body={
                "valueInputOption": "RAW",
                "data": data_to_update
            }
        ).execute()

        try:
            metadata = sheets_service.spreadsheets().get(spreadsheetId=file_id).execute()
            for sheet_info in metadata.get("sheets", []):
                if sheet_info["properties"]["title"] == "Sheet1":
                    default_sheet_id = sheet_info["properties"]["sheetId"]
                    sheets_service.spreadsheets().batchUpdate(
                        spreadsheetId=file_id,
                        body={"requests": [{"deleteSheet": {"sheetId": default_sheet_id}}]}
                    ).execute()
                    logging.info(f"Удален стандартный лист 'Sheet1' (ID: {default_sheet_id}).")
                    break
        except Exception as e:
            logging.warning(f"Не удалось удалить стандартный лист 'Sheet1': {e}")

        finalize_google_sheet_formatting(file_id, delete_columns_by_header=["RecommendedSort"])

        return {
            "success": True,
            "sheet_url": f"https://docs.google.com/spreadsheets/d/{file_id}"
        }

    except Exception as e:
        logging.error(f"Export failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))